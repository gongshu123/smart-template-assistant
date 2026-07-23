from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


FIELD_ALIASES = {
    "image": {"产品图", "产品图片", "图片", "素材", "图片文件"},
    "product_name": {"产品名称", "商品名称", "品名"},
    "selling_point": {"产品卖点", "商品卖点", "卖点", "大标题"},
    "price": {"价格", "到手价", "售价"},
    "benefit": {"利益点", "权益", "优惠利益点"},
    "discount": {"立减", "直降", "减"},
    "sales": {"销量", "已售", "销售量"},
}


@dataclass
class ProductData:
    image: str = ""
    product_name: str = ""
    selling_point: str = ""
    price: str = ""
    benefit: str = ""
    discount: str = ""
    sales: str = ""


@dataclass
class PageData:
    name: str
    products: list[ProductData] = field(default_factory=list)


@dataclass
class WorkbookData:
    source_path: str
    format_name: str
    pages: list[PageData]
    warnings: list[str] = field(default_factory=list)


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _field_key(label: Any) -> str | None:
    value = _clean(label).replace(" ", "")
    for field, aliases in FIELD_ALIASES.items():
        if value in aliases:
            return field
    return None


def _parse_transposed(sheet) -> WorkbookData:
    headers = [_clean(cell.value) for cell in sheet[1]]
    suite_columns = []
    for column in range(4, sheet.max_column + 1):
        header = headers[column - 1] or f"套版{column - 3:02d}"
        values = [_clean(sheet.cell(row=row, column=column).value) for row in range(2, sheet.max_row + 1)]
        meaningful = [value for value in values[:-1] if value]
        if meaningful:
            suite_columns.append((column, header))

    row_blocks: list[dict[str, int]] = []
    current: dict[str, int] | None = None
    for row in range(2, sheet.max_row + 1):
        layer_type = _clean(sheet.cell(row=row, column=1).value)
        layer_name = _clean(sheet.cell(row=row, column=2).value)
        if layer_type == "图片目录路径":
            break
        if layer_type == "图片":
            if current:
                row_blocks.append(current)
            current = {"image": row}
            continue
        if current and layer_type == "文本":
            key = _field_key(layer_name)
            if key:
                current[key] = row
    if current:
        row_blocks.append(current)

    pages: list[PageData] = []
    warnings: list[str] = []
    for column, header in suite_columns:
        products = []
        for slot_index, block in enumerate(row_blocks, start=1):
            values = {field: _clean(sheet.cell(row=row, column=column).value) for field, row in block.items()}
            product = ProductData(**{field: values.get(field, "") for field in ProductData.__dataclass_fields__})
            if any(vars(product).values()):
                products.append(product)
            else:
                warnings.append(f"{header} 的商品位{slot_index}为空")
        if products:
            pages.append(PageData(name=header, products=products))

    return WorkbookData(source_path="", format_name="横向套版表", pages=pages, warnings=warnings)


def _parse_rows(sheet, slot_count: int) -> WorkbookData:
    headers = [_clean(cell.value) for cell in sheet[1]]
    column_fields: dict[int, str] = {}
    page_column = None
    order_column = None
    output_column = None
    for index, header in enumerate(headers, start=1):
        normalized = header.replace(" ", "")
        field = _field_key(normalized)
        if field:
            column_fields[index] = field
        elif normalized in {"页面编号", "页码", "页面", "套版编号"}:
            page_column = index
        elif normalized in {"商品顺序", "顺序", "序号", "商品位"}:
            order_column = index
        elif normalized in {"输出文件名", "文件名", "输出名称"}:
            output_column = index

    rows: list[tuple[str, int, str, ProductData]] = []
    for row in range(2, sheet.max_row + 1):
        values = {field: _clean(sheet.cell(row=row, column=column).value) for column, field in column_fields.items()}
        product = ProductData(**{field: values.get(field, "") for field in ProductData.__dataclass_fields__})
        if not any(vars(product).values()):
            continue
        page_key = _clean(sheet.cell(row=row, column=page_column).value) if page_column else str((len(rows) // max(1, slot_count)) + 1)
        order_raw = _clean(sheet.cell(row=row, column=order_column).value) if order_column else ""
        try:
            order = int(float(order_raw)) if order_raw else len(rows) + 1
        except ValueError:
            order = len(rows) + 1
        output_name = _clean(sheet.cell(row=row, column=output_column).value) if output_column else ""
        rows.append((page_key or "1", order, output_name, product))

    grouped: dict[str, list[tuple[int, str, ProductData]]] = {}
    for page_key, order, output_name, product in rows:
        grouped.setdefault(page_key, []).append((order, output_name, product))

    pages = []
    warnings = []
    for page_key, items in grouped.items():
        items.sort(key=lambda item: item[0])
        name = next((item[1] for item in items if item[1]), f"套版{page_key}")
        products = [item[2] for item in items]
        if len(products) > slot_count:
            warnings.append(f"{name} 有{len(products)}个商品，但模板只有{slot_count}个商品位，多余数据将忽略")
        elif len(products) < slot_count:
            warnings.append(f"{name} 只有{len(products)}个商品，模板有{slot_count}个商品位")
        pages.append(PageData(name=name, products=products[:slot_count]))

    return WorkbookData(source_path="", format_name="标准逐行表", pages=pages, warnings=warnings)


def parse_workbook(path: str | Path, slot_count: int) -> WorkbookData:
    source = Path(path)
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    first_headers = {_clean(cell.value).replace(" ", "") for cell in sheet[1]}
    if {"图层类型", "图层名称", "变量名称"}.issubset(first_headers):
        result = _parse_transposed(sheet)
    else:
        result = _parse_rows(sheet, slot_count)
    result.source_path = str(source.resolve())
    workbook.close()
    return result


def validate_assets(data: WorkbookData, asset_folder: str | Path) -> list[str]:
    folder = Path(asset_folder)
    missing = []
    for page in data.pages:
        for index, product in enumerate(page.products, start=1):
            if product.image and not (folder / product.image).exists():
                missing.append(f"{page.name} 商品位{index}: {product.image}")
    return missing
